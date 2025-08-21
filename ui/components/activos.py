# ------ Imports ------ #
import tkinter as tk
from tkinter import ttk, messagebox
from tkinter.filedialog import asksaveasfilename

from database.connection import get_connection
from database.history import get_user_id_by_username, log_historial

from datetime import datetime, date

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from ui.components.ui_theme import FONT_LABEL, FONT_TITLE, BTN_KW

#-----------------------------------------------------------------

class ActivosWindow:
    def __init__(self, master, current_username: str, current_role: str):
        self.master = master
        master.title("Gestión de Activos")
        master.geometry("1300x800")

        self.current_username = current_username
        self.current_role = current_role
        self.current_user_id = get_user_id_by_username(current_username)

        # --- Filtros ---
        filter_bar = tk.Frame(master, padx=10, pady=8)
        filter_bar.pack(fill="x")

        # Categoría
        tk.Label(filter_bar, text="Categoría:").pack(side="left")
        self.cb_filter_cat = ttk.Combobox(filter_bar, width=34, state="readonly")
        self.cb_filter_cat.pack(side="left", padx=(6, 16))
        self._load_filter_categories()

        # Estado
        tk.Label(filter_bar, text="Estado:").pack(side="left")
        self.cb_filter_state = ttk.Combobox(filter_bar, width=12, state="readonly",
                                            values=["Todos", "Activo", "Baja"])
        self.cb_filter_state.current(0)
        self.cb_filter_state.pack(side="left", padx=(6, 16))

        # Texto
        tk.Label(filter_bar, text="Texto:").pack(side="left")
        self.e_filter_text = tk.Entry(filter_bar, width=32)
        self.e_filter_text.pack(side="left", padx=(6, 16))

        tk.Button(filter_bar, text="Aplicar", width=12, command=self.apply_filters).pack(side="left", padx=4)
        tk.Button(filter_bar, text="Limpiar", width=12, command=self.clear_filters).pack(side="left", padx=4)

        self.active_category_filter = None
        self.active_state_filter = None  # "Activo" | "Baja" | None
        self.active_text_filter = ""

        # --- Tabla ---
        self.tree = ttk.Treeview(master, columns=("id","nombre","categoria","ubicacion","estado"), show="headings")
        for col, txt, w in [
            ("id","ID",70),
            ("nombre","Nombre",320),
            ("categoria","Categoría",230),
            ("ubicacion","Ubicación",260),
            ("estado","Estado",80),
        ]:
            self.tree.heading(col, text=txt)
            self.tree.column(col, width=w, anchor="w")
        self.tree.pack(fill="both", expand=True, pady=8, padx=8)

        # --- Botonera ---
        btn_frame = tk.Frame(master)
        btn_frame.pack(pady=6)

        tk.Button(btn_frame, text="Agregar", width=16, height=2, command=self.add_activo).grid(row=0, column=0, padx=6)
        tk.Button(btn_frame, text="Modificar", width=16, height=2, command=self.edit_activo).grid(row=0, column=1, padx=6)

        col_idx = 2
        if self.current_role == "admin":
            tk.Button(btn_frame, text="Dar de Baja", width=16, height=2, command=self.baja_activo).grid(row=0, column=col_idx, padx=6)
            col_idx += 1
            tk.Button(btn_frame, text="Eliminar", width=16, height=2, command=self.delete_activo).grid(row=0, column=col_idx, padx=6);
            col_idx += 1

        tk.Button(btn_frame, text="Ver detalles", width=16, height=2, command=self.view_details).grid(row=0, column=col_idx, padx=6); col_idx += 1

        tk.Button(btn_frame, text="Editar detalles", width=16, height=2, command=self.edit_details).grid(row=0, column=col_idx, padx=6); col_idx += 1

        # Botón de Exportar
        tk.Button(btn_frame, text="Exportar Excel", width=20, height=2, command=self.export_excel).grid(row=0, column=col_idx, padx=6)
        col_idx += 1
        
        # Botón de Historial
        tk.Button(btn_frame, text="Historial", width=16, height=2, command=self.open_asset_history).grid(row=0, column=col_idx, padx=6)
        col_idx += 1

        tk.Button(btn_frame, text="Refrescar", width=16, height=2, command=self.load_activos).grid(row=0, column=col_idx, padx=6); col_idx += 1

        self.load_activos()

    # ---------------- Filtros ----------------
    def _load_filter_categories(self):
        conn = get_connection(); cur = conn.cursor()
        cur.execute("SELECT id_categoria, nombre FROM categorias_asset ORDER BY nombre")
        cats = cur.fetchall()
        conn.close()
        values = ["Todos"] + [f"{c[0]} - {c[1]}" for c in cats]
        self.cb_filter_cat["values"] = values
        self.cb_filter_cat.set("Todos")

    def apply_filters(self):
        # Categoría
        val = self.cb_filter_cat.get().strip()
        if not val or val == "Todos":
            self.active_category_filter = None
        else:
            try:
                self.active_category_filter = int(val.split(" - ")[0])
            except Exception:
                self.active_category_filter = None
        # Estado
        st = self.cb_filter_state.get().strip()
        if st in ("Activo", "Baja"):
            self.active_state_filter = st
        else:
            self.active_state_filter = None
        # Texto
        self.active_text_filter = self.e_filter_text.get().strip()
        self.load_activos()

    def clear_filters(self):
        self.cb_filter_cat.set("Todos")
        self.cb_filter_state.set("Todos")
        self.e_filter_text.delete(0, tk.END)
        self.active_category_filter = None
        self.active_state_filter = None
        self.active_text_filter = ""
        self.load_activos()

    # --------------- Carga de datos ---------------
    def load_activos(self):
        for it in self.tree.get_children():
            self.tree.delete(it)

        conn = get_connection(); cur = conn.cursor()
        sql = """
            SELECT a.id_asset, a.nombre, c.nombre, a.ubicacion,
                   CASE WHEN a.fecha_baja IS NULL OR a.fecha_baja='' THEN 'Activo' ELSE 'Baja' END AS estado
            FROM assets a
            LEFT JOIN categorias_asset c ON a.id_categoria = c.id_categoria
        """
        where = []
        params = []

        # Filtro por categoría
        if self.active_category_filter:
            where.append("a.id_categoria = ?")
            params.append(self.active_category_filter)

        # Filtro por estado
        if self.active_state_filter == "Activo":
            where.append("(a.fecha_baja IS NULL OR a.fecha_baja='')")
        elif self.active_state_filter == "Baja":
            where.append("(a.fecha_baja IS NOT NULL AND a.fecha_baja <> '')")

        # Filtro por texto (nombre, identificador, ubicación)
        if self.active_text_filter:
            where.append("(a.nombre LIKE ? OR a.identificador LIKE ? OR a.ubicacion LIKE ?)")
            wild = f"%{self.active_text_filter}%"
            params.extend([wild, wild, wild])

        if where:
            sql += " WHERE " + " AND ".join(where)
        sql += " ORDER BY a.id_asset"

        cur.execute(sql, params)
        rows = cur.fetchall()
        conn.close()

        for row in rows:
            self.tree.insert("", tk.END, values=row)

    # --------------- Utilidades ---------------
    def _get_selected(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showwarning("Atención", "Selecciona un activo.")
            return None
        return self.tree.item(sel[0])["values"]  # (id, nombre, categoria, ubicacion, estado)

    def _get_categoria_by_asset(self, id_asset: int):
        conn = get_connection(); cur = conn.cursor()
        cur.execute("""
            SELECT c.id_categoria, c.nombre
            FROM assets a
            LEFT JOIN categorias_asset c ON c.id_categoria = a.id_categoria
            WHERE a.id_asset=?
        """, (id_asset,))
        row = cur.fetchone()
        conn.close()
        return row  # (id_categoria, nombre)

    # --------------- Ver detalles (solo lectura) ---------------
    def view_details(self):
        selected = self._get_selected()
        if not selected: return
        id_asset, nombre, categoria, ubicacion, estado = selected

        details = []
        conn = get_connection(); cur = conn.cursor()
        cur.execute("""
            SELECT id_asset, nombre, identificador, fecha_alta, fecha_baja, ubicacion, observaciones
            FROM assets WHERE id_asset=?
        """, (id_asset,))
        gen = cur.fetchone()
        if gen:
            _, g_nombre, g_ident, g_falta, g_fbaja, g_ubi, g_obs = gen
            details.extend([
                ("Nombre", g_nombre or ""), ("Identificador", g_ident or ""),
                ("Fecha alta", g_falta or ""), ("Fecha baja", g_fbaja or ""),
                ("Ubicación", g_ubi or ""), ("Observaciones", g_obs or "")
            ])

        cat_name = (categoria or "").strip().lower()

        if "computadora" in cat_name:
            cur.execute("""
                SELECT ip, motherboard, procesador, ram_tipo, ram_cantidad, ssd, hdd, placa_red_ext
                FROM assets_computadoras WHERE id_asset=?
            """, (id_asset,))
            row = cur.fetchone()
            if row:
                details += [
                    ("IP", row[0] or ""), ("Motherboard", row[1] or ""), ("Procesador", row[2] or ""),
                    ("RAM Tipo", row[3] or ""), ("RAM Cantidad (GB)", "" if row[4] is None else str(row[4])),
                    ("SSD", row[5] or ""), ("HDD", row[6] or ""), ("Placa red externa", row[7] or "")
                ]

        elif "impresora" in cat_name:
            cur.execute("""
                SELECT marca, modelo, nro_serie, ip FROM assets_impresoras WHERE id_asset=?
            """, (id_asset,))
            row = cur.fetchone()
            if row:
                details += [("Marca", row[0] or ""), ("Modelo", row[1] or ""),
                            ("Nro. Serie", row[2] or ""), ("IP", row[3] or "")]

        elif "perifer" in cat_name:
            cur.execute("""
                SELECT tipo, marca, fecha_registro, fecha_entrega FROM assets_perifericos WHERE id_asset=?
            """, (id_asset,))
            row = cur.fetchone()
            if row:
                details += [("Tipo", row[0] or ""), ("Marca", row[1] or ""),
                            ("Fecha registro", row[2] or ""), ("Fecha entrega", row[3] or "")]
        conn.close()

        self._show_details_popup(id_asset, categoria, details)

    # Mostrar detalles popup

    def _show_details_popup(self, id_asset: int, categoria: str, details: list[tuple[str, str]]):
        pop = tk.Toplevel(self.master)
        pop.title(f"Detalles del Activo #{id_asset} · {categoria}")
        pop.geometry("620x560")
        pop.resizable(False, False)

        wrap = tk.Frame(pop, padx=12, pady=12)
        wrap.pack(fill="both", expand=True)

        canvas = tk.Canvas(wrap)
        vsb = tk.Scrollbar(wrap, orient="vertical", command=canvas.yview)
        frm = tk.Frame(canvas)

        frm.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0,0), window=frm, anchor="nw")
        canvas.configure(yscrollcommand=vsb.set)

        canvas.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")

        for label, val in details:
            row = tk.Frame(frm)
            row.pack(fill="x", pady=3)
            tk.Label(row, text=f"{label}:", width=22, anchor="w").pack(side="left")
            tk.Label(row, text=val, anchor="w", justify="left", wraplength=420).pack(side="left", fill="x", expand=True)

        tk.Button(frm, text="Cerrar", width=14, command=pop.destroy).pack(pady=12)

    # --------------- ABM ---------------
    def add_activo(self):
        popup = tk.Toplevel(self.master)
        popup.title("Nuevo Activo")
        popup.geometry("520x430")

        tk.Label(popup, text="Nombre:").pack(pady=5, anchor="w")
        e_nombre = tk.Entry(popup, width=52); e_nombre.pack()

        tk.Label(popup, text="Identificador:").pack(pady=5, anchor="w")
        e_ident = tk.Entry(popup, width=52); e_ident.pack()

        tk.Label(popup, text="Ubicación:").pack(pady=5, anchor="w")
        e_ubi = tk.Entry(popup, width=52); e_ubi.pack()

        tk.Label(popup, text="Observaciones:").pack(pady=5, anchor="w")
        e_obs = tk.Entry(popup, width=52); e_obs.pack()

        tk.Label(popup, text="Categoría:").pack(pady=5, anchor="w")
        cb_cat = ttk.Combobox(popup, width=50); cb_cat.pack()

        conn = get_connection(); cur = conn.cursor()
        cur.execute("SELECT id_categoria, nombre FROM categorias_asset ORDER BY nombre")
        categorias = cur.fetchall()
        conn.close()
        cb_cat["values"] = [f"{c[0]} - {c[1]}" for c in categorias]

        def save():
            nombre = e_nombre.get().strip()
            cat_value = cb_cat.get().strip()
            if not nombre or not cat_value:
                messagebox.showerror("Error", "Nombre y Categoría son obligatorios")
                return

            identificador = e_ident.get().strip()
            ubicacion = e_ubi.get().strip()
            obs = e_obs.get().strip()
            id_categoria = int(cat_value.split(" - ")[0])
            fecha_alta = date.today().isoformat()

            conn = get_connection(); cur = conn.cursor()
            cur.execute("""
                INSERT INTO assets (id_categoria, nombre, identificador, fecha_alta, ubicacion, observaciones)
                VALUES (?, ?, ?, ?, ?, ?)
            """, (id_categoria, nombre, identificador, fecha_alta, ubicacion, obs))
            conn.commit()
            new_id = cur.lastrowid

            # Traer nombre de categoría (para ruteo de detalles)
            cur.execute("SELECT nombre FROM categorias_asset WHERE id_categoria=?", (id_categoria,))
            cat_row = cur.fetchone()
            conn.close()

            if self.current_user_id:
                log_historial(new_id, self.current_user_id, f"Creación del activo '{nombre}' (cat={id_categoria})")

            # ⬇️ NUEVO: abrir editor de detalles según categoría
            categoria_nombre = (cat_row[0] if cat_row else "").strip()
            popup.destroy()
            self._open_details_editor(new_id, categoria_nombre)  # abre el popup de detalles
            self.load_activos()

        tk.Button(popup, text="Guardar", width=18, height=2, command=save).pack(pady=12)

    # Editar Activo

    def edit_activo(self):
        selected = self._get_selected()
        if not selected: return
        id_asset, nombre, categoria, ubicacion, estado = selected

        popup = tk.Toplevel(self.master)
        popup.title("Editar Activo")
        popup.geometry("520x320")

        tk.Label(popup, text="Nombre:").pack(pady=5, anchor="w")
        e_nombre = tk.Entry(popup, width=52); e_nombre.insert(0, nombre); e_nombre.pack()

        tk.Label(popup, text="Ubicación:").pack(pady=5, anchor="w")
        e_ubi = tk.Entry(popup, width=52); e_ubi.insert(0, ubicacion); e_ubi.pack()

        tk.Label(popup, text="Observaciones:").pack(pady=5, anchor="w")
        e_obs = tk.Entry(popup, width=52); e_obs.pack()

        def save_changes():
            new_nombre = e_nombre.get().strip()
            new_ubi = e_ubi.get().strip()
            new_obs = e_obs.get().strip()

            conn = get_connection(); cur = conn.cursor()
            cur.execute("UPDATE assets SET nombre=?, ubicacion=?, observaciones=? WHERE id_asset=?",
                        (new_nombre, new_ubi, new_obs, id_asset))
            conn.commit()
            conn.close()

            if self.current_user_id:
                log_historial(id_asset, self.current_user_id,
                              f"Modificación: nombre='{nombre}'→'{new_nombre}', ubicación='{ubicacion}'→'{new_ubi}'")

            popup.destroy()
            self.load_activos()
            messagebox.showinfo("Éxito", "Activo modificado.")

        tk.Button(popup, text="Guardar cambios", width=18, height=2, command=save_changes).pack(pady=12)

    # Dar de baja activo

    def baja_activo(self):
        selected = self._get_selected()
        if not selected: return
        id_asset, nombre, categoria, ubicacion, estado = selected
        if estado == "Baja":
            messagebox.showinfo("Info", "El activo ya está dado de baja.")
            return

        if not messagebox.askyesno("Confirmar", f"¿Dar de baja el activo {id_asset} - '{nombre}'?"):
            return

        fecha_baja = datetime.date.today().isoformat()
        conn = get_connection(); cur = conn.cursor()
        cur.execute("UPDATE assets SET fecha_baja=? WHERE id_asset=?", (fecha_baja, id_asset))
        conn.commit(); conn.close()

        if self.current_user_id:
            log_historial(id_asset, self.current_user_id, f"Baja del activo '{nombre}' (fecha_baja={fecha_baja})")

        self.load_activos()
        messagebox.showinfo("Éxito", "Activo dado de baja.")

    # Delete Activo

    def delete_activo(self):
        selected = self._get_selected()
        if not selected: return
        id_asset, nombre, categoria, ubicacion, estado = selected

        if not messagebox.askyesno("Confirmar", f"¿Eliminar DEFINITIVAMENTE el activo {id_asset} - '{nombre}'?"):
            return

        conn = get_connection(); cur = conn.cursor()
        try:
            # Desasociar historial (si la FK no tiene ON DELETE SET NULL)
            cur.execute("UPDATE historial_cambios SET id_asset=NULL WHERE id_asset=?", (id_asset,))
            # Borrar detalles específicos primero
            cur.execute("DELETE FROM assets_computadoras WHERE id_asset=?", (id_asset,))
            cur.execute("DELETE FROM assets_impresoras   WHERE id_asset=?", (id_asset,))
            cur.execute("DELETE FROM assets_perifericos  WHERE id_asset=?", (id_asset,))
            # Borrar asset
            cur.execute("DELETE FROM assets WHERE id_asset=?", (id_asset,))
            conn.commit()
        except Exception as e:
            conn.rollback()
            messagebox.showerror("Error", f"No se pudo eliminar: {e}")
            return
        finally:
            conn.close()

        if self.current_user_id:
            log_historial(None, self.current_user_id, f"Eliminación definitiva del activo '{nombre}' (id={id_asset})")

        self.load_activos()
        messagebox.showinfo("Éxito", "Activo eliminado.")

    # Exportar Excel 

    def _fetch_filtered_assets(self):
        """
        Devuelve:
        - listed_rows: (id_asset, nombre, categoria, ubicacion, estado) según filtros actuales
        - base_rows:   (id_asset, id_categoria, categoria, nombre, identificador, fecha_alta, fecha_baja, ubicacion, observaciones)
        - comp_map:    {id_asset: (ip, motherboard, procesador, ram_tipo, ram_cantidad, ssd, hdd, placa_red_ext)}
        - imp_map:     {id_asset: (marca, modelo, nro_serie, ip)}
        - per_map:     {id_asset: (tipo, marca, fecha_registro, fecha_entrega)}
        """
        conn = get_connection(); cur = conn.cursor()

        # 1) listado filtrado (igual que load_activos)
        sql = """
            SELECT a.id_asset, a.nombre, c.nombre AS categoria, a.ubicacion,
                CASE WHEN a.fecha_baja IS NULL OR a.fecha_baja='' THEN 'Activo' ELSE 'Baja' END AS estado
            FROM assets a
            LEFT JOIN categorias_asset c ON a.id_categoria = c.id_categoria
        """
        where, params = [], []

        if getattr(self, "active_category_filter", None):
            where.append("a.id_categoria = ?")
            params.append(self.active_category_filter)

        if getattr(self, "active_state_filter", None) == "Activo":
            where.append("(a.fecha_baja IS NULL OR a.fecha_baja='')")
        elif getattr(self, "active_state_filter", None) == "Baja":
            where.append("(a.fecha_baja IS NOT NULL AND a.fecha_baja <> '')")

        txt = getattr(self, "active_text_filter", "").strip()
        if txt:
            like = f"%{txt}%"
            where.append("(a.nombre LIKE ? OR a.identificador LIKE ? OR a.ubicacion LIKE ?)")
            params.extend([like, like, like])

        if where:
            sql += " WHERE " + " AND ".join(where)
        sql += " ORDER BY a.id_asset"

        cur.execute(sql, params)
        listed_rows = cur.fetchall()

        # 2) base + detalles para esos IDs
        ids = [r[0] for r in listed_rows]
        base_rows, comp_map, imp_map, per_map = [], {}, {}, {}

        if ids:
            qmarks = ",".join(["?"]*len(ids))

            cur.execute(f"""
                SELECT a.id_asset, a.id_categoria, c.nombre AS categoria, a.nombre, a.identificador,
                    a.fecha_alta, a.fecha_baja, a.ubicacion, a.observaciones
                FROM assets a
                LEFT JOIN categorias_asset c ON c.id_categoria = a.id_categoria
                WHERE a.id_asset IN ({qmarks})
                ORDER BY a.id_asset
            """, ids)
            base_rows = cur.fetchall()

            cur.execute(f"SELECT id_asset, ip, motherboard, procesador, ram_tipo, ram_cantidad, ssd, hdd, placa_red_ext FROM assets_computadoras WHERE id_asset IN ({qmarks})", ids)
            for r in cur.fetchall():
                comp_map[r[0]] = r[1:]

            cur.execute(f"SELECT id_asset, marca, modelo, nro_serie, ip FROM assets_impresoras WHERE id_asset IN ({qmarks})", ids)
            for r in cur.fetchall():
                imp_map[r[0]] = r[1:]

            cur.execute(f"SELECT id_asset, tipo, marca, fecha_registro, fecha_entrega FROM assets_perifericos WHERE id_asset IN ({qmarks})", ids)
            for r in cur.fetchall():
                per_map[r[0]] = r[1:]

        conn.close()
        return listed_rows, base_rows, comp_map, imp_map, per_map

       
    def export_excel(self):
        # 1) Tomar filtros y datos de la vista actual
        listed_rows, base_rows, comp_map, imp_map, per_map = self._fetch_filtered_assets()
        if not listed_rows:
            messagebox.showinfo("Exportar", "No hay datos para exportar con los filtros actuales.")
            return

        # 2) Categorías presentes en el resultado
        categorias_presentes = sorted({ (r[2] or "").strip() for r in listed_rows if r[2] })
        if len(categorias_presentes) > 1:
            msg = (
                "No se puede exportar múltiples categorías a la vez "
                "(p. ej. Computadoras e Impresoras) por incompatibilidad de detalles.\n\n"
                "Aplicá un filtro de ‘Categoría’ para una sola categoría e intentá nuevamente."
            )
            messagebox.showerror("Exportar - Incompatible", msg)
            return

        categoria_unica = categorias_presentes[0] if categorias_presentes else "Sin categoría"
        cat_name_l = categoria_unica.lower()

        # 3) Definir columnas base + específicas según la categoría
        base_headers = [
            "id_asset", "categoria", "nombre", "identificador",
            "fecha_alta", "fecha_baja", "ubicacion", "observaciones", "estado"
        ]
        if "computadora" in cat_name_l:
            detail_headers = ["ip", "motherboard", "procesador", "ram_tipo", "ram_cantidad", "ssd", "hdd", "placa_red_ext"]
            detail_map = comp_map
            sheet_title = "Computadoras"
        elif "impresora" in cat_name_l:
            detail_headers = ["marca", "modelo", "nro_serie", "ip"]
            detail_map = imp_map
            sheet_title = "Impresoras"
        elif "perifer" in cat_name_l:  # periférico
            detail_headers = ["tipo", "marca", "fecha_registro", "fecha_entrega"]
            detail_map = per_map
            sheet_title = "Perifericos"
        else:
            # Categoría sin tabla específica -> solo base
            detail_headers = []
            detail_map = {}
            sheet_title = categoria_unica or "Assets"

        # 4) Construir filas: solo los IDs que están en el listado filtrado
        ids_listados = [r[0] for r in listed_rows]
        base_map = {r[0]: r for r in base_rows}  # (id_asset, id_categoria, categoria, nombre, identificador, fecha_alta, fecha_baja, ubicacion, observaciones)

        def estado_from_base(row):
            # fecha_baja en pos 6
            return "Activo" if (row[6] is None or row[6] == "") else "Baja"

        headers = base_headers + detail_headers
        rows_out = []
        for aid in ids_listados:
            b = base_map.get(aid)
            if not b:
                continue
            base_vals = [b[0], b[2], b[3], b[4], b[5], b[6], b[7], b[8], estado_from_base(b)]
            det_vals = list(detail_map.get(aid, [None]*len(detail_headers))) if detail_headers else []
            rows_out.append(base_vals + det_vals)

        # 5) Crear workbook con una sola hoja (la de la categoría)
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_title
        ws.append(headers)
        for row in rows_out:
            ws.append(row)

        # formato
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}1"
        for col in range(1, ws.max_column+1):
            ws.column_dimensions[get_column_letter(col)].width = 20

        # 6) Guardar
        ts = datetime.now().strftime("%Y%m%d_%H%M")
        default_name = f"Export_{sheet_title}_{ts}.xlsx"
        path = asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile=default_name,
            title="Guardar exportación por categoría"
        )
        if not path:
            return

        try:
            wb.save(path)
            messagebox.showinfo("Exportar", f"Exportación creada:\n{path}")
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo guardar el Excel:\n{e}")

    # Historial
    
    def open_asset_history(self):
        sel = self._get_selected()
        if not sel:
            return
        id_asset, nombre, categoria, *_ = sel

        pop = tk.Toplevel(self.master)
        pop.title(f"Historial · Activo #{id_asset} - {nombre}")
        pop.geometry("760x420")
        pop.resizable(False, False)

        cols = ("fecha","usuario","descripcion")
        tree = ttk.Treeview(pop, columns=cols, show="headings")
        for col, txt, w in [("fecha","Fecha",160), ("usuario","Usuario",180), ("descripcion","Descripción",380)]:
            tree.heading(col, text=txt); tree.column(col, width=w, anchor="w")
        tree.pack(fill="both", expand=True, padx=8, pady=8)

        conn = get_connection(); cur = conn.cursor()
        cur.execute("""
            SELECT h.fecha_cambio, COALESCE(u.nombre, '-'), h.descripcion_cambio
            FROM historial_cambios h
            LEFT JOIN usuarios u ON u.id_usuario = h.id_usuario
            WHERE h.id_asset = ?
            ORDER BY h.id_historial DESC
        """, (id_asset,))
        for row in cur.fetchall():
            tree.insert("", tk.END, values=row)
        conn.close()

        tk.Button(pop, text="Cerrar", command=pop.destroy, **BTN_KW).pack(pady=(0,8))

    def edit_details(self):
        sel = self._get_selected()
        if not sel:
            return
        id_asset, nombre, categoria, *_ = sel
        self._open_details_editor(id_asset, categoria)

    def _open_details_editor(self, id_asset: int, categoria_nombre: str):
        cat = (categoria_nombre or "").lower()
        if "computadora" in cat:
            self._edit_computadora_details(id_asset)
        elif "impresora" in cat:
            self._edit_impresora_details(id_asset)
        elif "perifer" in cat:  # periférico
            self._edit_periferico_details(id_asset)
        else:
            messagebox.showinfo("Detalles", "Esta categoría no tiene detalles específicos.")
            return

# ---------- COMPUTADORA ----------
    def _edit_computadora_details(self, id_asset: int):
        pop = tk.Toplevel(self.master)
        pop.title(f"Detalles · Computadora (Asset #{id_asset})")
        pop.geometry("520x420")
        pop.resizable(False, False)

        labels = ["IP","Motherboard","Procesador","RAM Tipo","RAM Cantidad (GB)","SSD","HDD","Placa red externa"]
        entries = []
        for i, lbl in enumerate(labels):
            tk.Label(pop, text=lbl+":").pack(anchor="w", padx=10, pady=(8 if i==0 else 4, 0))
            e = tk.Entry(pop, width=52); e.pack(padx=10)
            entries.append(e)

        # Cargar existentes
        conn = get_connection(); cur = conn.cursor()
        cur.execute("""
            SELECT ip, motherboard, procesador, ram_tipo, ram_cantidad, ssd, hdd, placa_red_ext
            FROM assets_computadoras WHERE id_asset=?
        """, (id_asset,))
        row = cur.fetchone()
        if row:
            for e, v in zip(entries, row):
                e.insert(0, "" if v is None else str(v))
        conn.close()

        def save():
            vals = [e.get().strip() or None for e in entries]
            # ram_cantidad a int o None
            if vals[4] is not None:
                try:
                    vals[4] = int(vals[4])
                except ValueError:
                    messagebox.showerror("Error", "RAM Cantidad debe ser un número.")
                    return
            conn2 = get_connection(); c2 = conn2.cursor()
            # upsert manual
            c2.execute("SELECT 1 FROM assets_computadoras WHERE id_asset=?", (id_asset,))
            exists = c2.fetchone() is not None
            if exists:
                c2.execute("""
                    UPDATE assets_computadoras
                    SET ip=?, motherboard=?, procesador=?, ram_tipo=?, ram_cantidad=?, ssd=?, hdd=?, placa_red_ext=?
                    WHERE id_asset=?
                """, (*vals, id_asset))
            else:
                c2.execute("""
                    INSERT INTO assets_computadoras
                        (id_asset, ip, motherboard, procesador, ram_tipo, ram_cantidad, ssd, hdd, placa_red_ext)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (id_asset, *vals))
            conn2.commit(); conn2.close()
            if self.current_user_id:
                log_historial(id_asset, self.current_user_id, "Actualización de detalles (Computadora)")
            pop.destroy()
            self.load_activos()
            messagebox.showinfo("Éxito", "Detalles guardados.")

        tk.Button(pop, text="Guardar", width=16, height=2, command=save).pack(pady=12)

# ---------- IMPRESORA ----------
    def _edit_impresora_details(self, id_asset: int):
        pop = tk.Toplevel(self.master)
        pop.title(f"Detalles · Impresora (Asset #{id_asset})")
        pop.geometry("480x360")
        pop.resizable(False, False)

        fields = [("Marca",""), ("Modelo",""), ("Nro. Serie",""), ("IP","")]
        entries = []
        for i, (lbl, _) in enumerate(fields):
            tk.Label(pop, text=lbl+":").pack(anchor="w", padx=10, pady=(8 if i==0 else 4, 0))
            e = tk.Entry(pop, width=48); e.pack(padx=10)
            entries.append(e)

        # Cargar existentes
        conn = get_connection(); cur = conn.cursor()
        cur.execute("SELECT marca, modelo, nro_serie, ip FROM assets_impresoras WHERE id_asset=?", (id_asset,))
        row = cur.fetchone()
        if row:
            for e, v in zip(entries, row):
                e.insert(0, "" if v is None else str(v))
        conn.close()

        def save():
            vals = [e.get().strip() or None for e in entries]
            conn2 = get_connection(); c2 = conn2.cursor()
            c2.execute("SELECT 1 FROM assets_impresoras WHERE id_asset=?", (id_asset,))
            exists = c2.fetchone() is not None
            if exists:
                c2.execute("""
                    UPDATE assets_impresoras
                    SET marca=?, modelo=?, nro_serie=?, ip=?
                    WHERE id_asset=?
                """, (*vals, id_asset))
            else:
                c2.execute("""
                    INSERT INTO assets_impresoras (id_asset, marca, modelo, nro_serie, ip)
                    VALUES (?, ?, ?, ?, ?)
                """, (id_asset, *vals))
            conn2.commit(); conn2.close()
            if self.current_user_id:
                log_historial(id_asset, self.current_user_id, "Actualización de detalles (Impresora)")
            pop.destroy()
            self.load_activos()
            messagebox.showinfo("Éxito", "Detalles guardados.")

        tk.Button(pop, text="Guardar", width=16, height=2, command=save).pack(pady=12)

# ---------- PERIFÉRICO ----------
    def _edit_periferico_details(self, id_asset: int):
        pop = tk.Toplevel(self.master)
        pop.title(f"Detalles · Periférico (Asset #{id_asset})")
        pop.geometry("520x380")
        pop.resizable(False, False)

        # Campos
        tk.Label(pop, text="Tipo:").pack(anchor="w", padx=10, pady=(8, 0))
        e_tipo = tk.Entry(pop, width=52); e_tipo.pack(padx=10)

        tk.Label(pop, text="Marca:").pack(anchor="w", padx=10, pady=(8, 0))
        e_marca = tk.Entry(pop, width=52); e_marca.pack(padx=10)

        tk.Label(pop, text="Fecha registro (YYYY-MM-DD):").pack(anchor="w", padx=10, pady=(8, 0))
        e_freg = tk.Entry(pop, width=52); e_freg.pack(padx=10)

        tk.Label(pop, text="Fecha entrega (YYYY-MM-DD):").pack(anchor="w", padx=10, pady=(8, 0))
        e_fent = tk.Entry(pop, width=52); e_fent.pack(padx=10)

        # Cargar existentes / defaults
        conn = get_connection(); cur = conn.cursor()
        cur.execute("""
            SELECT tipo, marca, fecha_registro, fecha_entrega
            FROM assets_perifericos WHERE id_asset=?
        """, (id_asset,))
        row = cur.fetchone()
        if row:
            for e, v in zip((e_tipo, e_marca, e_freg, e_fent), row):
                e.insert(0, "" if v is None else str(v))
        else:
            # Defaults según definición (si no existen registros aún)
            cur.execute("SELECT fecha_alta FROM assets WHERE id_asset=?", (id_asset,))
            base = cur.fetchone()
            default_date = (base[0] if base and base[0] else date.today().isoformat())
            e_freg.insert(0, default_date); e_fent.insert(0, default_date)
        conn.close()

        def _valid_date_or_none(s):
            s = (s or "").strip()
            if not s:
                return None
            try:
                # validación simple
                _ = s.split("-")
                if len(_) != 3: raise ValueError
                return s
            except Exception:
                return None

        def save():
            tipo = e_tipo.get().strip() or None
            marca = e_marca.get().strip() or None
            f_reg = _valid_date_or_none(e_freg.get())
            f_ent = _valid_date_or_none(e_fent.get())
            if e_freg.get().strip() and not f_reg:
                messagebox.showerror("Error", "Fecha registro inválida. Use YYYY-MM-DD.")
                return
            if e_fent.get().strip() and not f_ent:
                messagebox.showerror("Error", "Fecha entrega inválida. Use YYYY-MM-DD.")
                return

            conn2 = get_connection(); c2 = conn2.cursor()
            c2.execute("SELECT 1 FROM assets_perifericos WHERE id_asset=?", (id_asset,))
            exists = c2.fetchone() is not None
            if exists:
                c2.execute("""
                    UPDATE assets_perifericos
                    SET tipo=?, marca=?, fecha_registro=?, fecha_entrega=?
                    WHERE id_asset=?
                """, (tipo, marca, f_reg, f_ent, id_asset))
            else:
                c2.execute("""
                    INSERT INTO assets_perifericos (id_asset, tipo, marca, fecha_registro, fecha_entrega)
                    VALUES (?, ?, ?, ?, ?)
                """, (id_asset, tipo, marca, f_reg, f_ent))
            conn2.commit(); conn2.close()
            if self.current_user_id:
                log_historial(id_asset, self.current_user_id, "Actualización de detalles (Periférico)")
            pop.destroy()
            self.load_activos()
            messagebox.showinfo("Éxito", "Detalles guardados.")

        tk.Button(pop, text="Guardar", width=16, height=2, command=save).pack(pady=12)
